"""Build the multi-sheet ``.xlsx`` report with openpyxl.

Six sheets:

* **Listing** – one row per active listing.
* **Novità** – listings whose ``first_seen`` falls inside the current run.
* **Variazioni di prezzo** – listings with ≥ 2 price-history entries.
* **Riepilogo** – KPIs of the current run + distributions and quantiles.
* **Run history** – one row per past run.
* **Configurazione** – dump of the YAML config + any errors encountered.

The whole module deliberately uses ``openpyxl`` formulas (not
hard-coded numbers) for €/m² and price quantiles so the user can edit the
spreadsheet manually and see numbers update.
"""

from __future__ import annotations

import logging
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from .models import listing_url

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------- styling
HEADER_FILL = PatternFill("solid", fgColor="005E69")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="center", wrap_text=False)

LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")

STALE_FONT = Font(name="Calibri", size=10, color="888888", italic=True)

RESTORE_FILL = PatternFill("solid", fgColor="FFF4CC")    # "Da ristrutturare"
AUCTION_FILL = PatternFill("solid", fgColor="EEEEEE")    # asta

POSITIVE_FONT = Font(name="Calibri", size=10, color="C00000")  # rincaro = rosso
NEGATIVE_FONT = Font(name="Calibri", size=10, color="006100")  # ribasso = verde

PRICE_FORMAT = '€#,##0;[Red]-€#,##0;"-"'


@dataclass
class ExportContext:
    """All the inputs the exporter needs in addition to the DB."""

    config_yaml: str
    run_started_at: datetime
    run_finished_at: datetime
    run_counters: dict[str, int]
    errors: list[dict[str, Any]] = field(default_factory=list)
    novita_since: datetime | None = None  # overrides run_started_at if set


# ---------------------------------------------------------------------- API


def export_workbook(
    db,
    output_path: str | Path,
    ctx: ExportContext,
) -> Path:
    """Build the workbook and write it to ``output_path``."""
    wb = Workbook()
    # openpyxl always creates an initial sheet — repurpose it.
    ws_listing = wb.active
    ws_listing.title = "Listing"

    listings_rows = db.all_active_listings()
    n_listing_rows = _build_listing_sheet(ws_listing, listings_rows)

    ws_novita = wb.create_sheet("Novità")
    novita_since = ctx.novita_since or ctx.run_started_at
    novita_rows = db.listings_first_seen_after(novita_since)
    _build_novita_sheet(ws_novita, novita_rows)

    ws_var = wb.create_sheet("Variazioni di prezzo")
    var_rows = db.listings_with_price_changes()
    _build_variazioni_sheet(ws_var, var_rows)

    ws_riepilogo = wb.create_sheet("Riepilogo")
    _build_riepilogo_sheet(
        wb,
        ws_riepilogo,
        ctx=ctx,
        active_count=db.count_active(),
        listings_rows=listings_rows,
        n_listing_rows=n_listing_rows,
    )

    ws_runs = wb.create_sheet("Run history")
    _build_run_history_sheet(ws_runs, db.list_runs())

    ws_cfg = wb.create_sheet("Configurazione")
    _build_config_sheet(ws_cfg, ctx)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ---------------------------------------------------------- "Listing" sheet


# Column layout shared between "Listing" and "Novità".
_LISTING_HEADERS = [
    "#", "ID", "Provincia", "Comune", "Tipologia", "Titolo", "Prezzo (€)",
    "Superficie (m²)", "€/m²", "Locali", "Bagni", "Stato",
    "Trattativa privata?", "Asta?", "Indirizzo", "Lat", "Lng",
    "First seen", "Last seen", "Variazione prezzo", "Status",
    "Ricerca", "Link",
]
# 0-based indices used when writing values
_COL_IDX = {name: i for i, name in enumerate(_LISTING_HEADERS)}
_COL = {name: i + 1 for i, name in enumerate(_LISTING_HEADERS)}  # 1-based


def _build_listing_sheet(ws, rows: list[sqlite3.Row]) -> int:
    """Fill the "Listing" sheet. Returns the number of data rows written."""
    _write_header(ws, _LISTING_HEADERS)

    # Pre-compute price variation strings using the price_history view from
    # the DB. We do it outside the loop to avoid per-row SQL.
    n_data = 0
    for ordinal, row in enumerate(rows, start=1):
        excel_row = ordinal + 1  # +1 because of header
        _write_listing_row(ws, excel_row, row, ordinal=ordinal)
        n_data += 1

    _finalize_listing_sheet(ws, n_data)
    return n_data


def _build_novita_sheet(ws, rows: list[sqlite3.Row]) -> None:
    if not rows:
        ws["A1"] = "Nessuna novità in questo run"
        ws["A1"].font = Font(name="Calibri", size=12, italic=True)
        return

    _write_header(ws, _LISTING_HEADERS)
    for ordinal, row in enumerate(rows, start=1):
        _write_listing_row(ws, ordinal + 1, row, ordinal=ordinal)
    _finalize_listing_sheet(ws, len(rows))


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 32


def _write_listing_row(ws, excel_row: int, row: sqlite3.Row, *, ordinal: int) -> None:
    listing_id = int(row["id"])
    surface_mq = row["superficie_mq"]
    prezzo = row["prezzo_corrente"]

    values: list[Any] = [None] * len(_LISTING_HEADERS)
    values[_COL_IDX["#"]] = ordinal
    values[_COL_IDX["ID"]] = listing_id
    values[_COL_IDX["Provincia"]] = row["provincia"] or ""
    values[_COL_IDX["Comune"]] = row["comune"] or ""
    values[_COL_IDX["Tipologia"]] = row["tipologia"] or ""
    values[_COL_IDX["Titolo"]] = row["titolo"] or ""
    values[_COL_IDX["Prezzo (€)"]] = prezzo
    # Numeric integer so Excel can sort/filter/aggregate it. If the raw
    # surface couldn't be parsed (e.g. "n.d.", missing field) ``surface_mq``
    # is None, which openpyxl renders as an empty cell — preferable to
    # writing the literal string "None" or echoing the raw text.
    values[_COL_IDX["Superficie (m²)"]] = surface_mq
    # €/m² uses an Excel formula so the user can edit prices and see it
    # update. ``IFERROR`` swallows the case where surface is missing.
    price_cell = f"{get_column_letter(_COL['Prezzo (€)'])}{excel_row}"
    values[_COL_IDX["€/m²"]] = (
        f"=IFERROR({price_cell}/{surface_mq},0)" if surface_mq else 0
    )
    values[_COL_IDX["Locali"]] = row["locali"]
    values[_COL_IDX["Bagni"]] = row["bagni"]
    values[_COL_IDX["Stato"]] = row["stato"] or ""
    values[_COL_IDX["Trattativa privata?"]] = bool(row["e_trattativa"])
    values[_COL_IDX["Asta?"]] = bool(row["e_asta"])
    values[_COL_IDX["Indirizzo"]] = row["indirizzo"] or ""
    values[_COL_IDX["Lat"]] = row["lat"]
    values[_COL_IDX["Lng"]] = row["lng"]
    values[_COL_IDX["First seen"]] = _fmt_dt(row["first_seen"])
    values[_COL_IDX["Last seen"]] = _fmt_dt(row["last_seen"])
    values[_COL_IDX["Variazione prezzo"]] = ""  # filled below if applicable
    values[_COL_IDX["Status"]] = "attivo" if row["attivo"] else "non più disponibile"
    values[_COL_IDX["Ricerca"]] = row["ricerca_nome"] or ""
    url = listing_url(listing_id)
    values[_COL_IDX["Link"]] = f'=HYPERLINK("{url}","Apri annuncio")'

    for col_off, val in enumerate(values, start=1):
        c = ws.cell(row=excel_row, column=col_off, value=val)
        c.font = BODY_FONT
        c.alignment = BODY_ALIGN

    # Type-specific formatting + highlights.
    ws.cell(row=excel_row, column=_COL["Prezzo (€)"]).number_format = PRICE_FORMAT
    ws.cell(row=excel_row, column=_COL["Superficie (m²)"]).number_format = "#,##0"
    ws.cell(row=excel_row, column=_COL["€/m²"]).number_format = PRICE_FORMAT
    ws.cell(row=excel_row, column=_COL["Link"]).font = LINK_FONT

    if (row["stato"] or "").strip().lower() == "da ristrutturare":
        _fill_row(ws, excel_row, len(_LISTING_HEADERS), RESTORE_FILL)
    if bool(row["e_asta"]):
        _fill_row(ws, excel_row, len(_LISTING_HEADERS), AUCTION_FILL)
    if not row["attivo"]:
        for col in range(1, len(_LISTING_HEADERS) + 1):
            ws.cell(row=excel_row, column=col).font = STALE_FONT


def _finalize_listing_sheet(ws, n_data: int) -> None:
    last_col_letter = get_column_letter(len(_LISTING_HEADERS))
    if n_data >= 1:
        ws.auto_filter.ref = f"A1:{last_col_letter}{n_data + 1}"
    ws.freeze_panes = "A2"

    widths = {
        "#": 5, "ID": 12, "Provincia": 10, "Comune": 18, "Tipologia": 16,
        "Titolo": 50, "Prezzo (€)": 14, "Superficie (m²)": 14, "€/m²": 12,
        "Locali": 8, "Bagni": 8, "Stato": 22, "Trattativa privata?": 10,
        "Asta?": 8, "Indirizzo": 30, "Lat": 10, "Lng": 10,
        "First seen": 18, "Last seen": 18, "Variazione prezzo": 18,
        "Status": 18, "Ricerca": 22, "Link": 16,
    }
    for name, width in widths.items():
        ws.column_dimensions[get_column_letter(_COL[name])].width = width


def _fill_row(ws, row: int, n_cols: int, fill: PatternFill) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = fill


# --------------------------------------------------- "Variazioni" sheet


def _build_variazioni_sheet(ws, rows: list[dict[str, Any]]) -> None:
    headers = [
        "ID", "Comune", "Titolo", "Prezzo iniziale", "Prezzo attuale",
        "Variazione €", "Variazione %", "Data prima", "Data ultima", "Link",
    ]
    _write_header(ws, headers)

    if not rows:
        ws["A2"] = "Nessuna variazione di prezzo registrata"
        ws["A2"].font = Font(name="Calibri", size=11, italic=True)
        return

    for i, r in enumerate(rows, start=2):
        url = listing_url(int(r["id"]))
        ws.cell(row=i, column=1, value=int(r["id"]))
        ws.cell(row=i, column=2, value=r["comune"] or "")
        ws.cell(row=i, column=3, value=r["titolo"] or "")
        ws.cell(row=i, column=4, value=r["prezzo_iniziale"]).number_format = PRICE_FORMAT
        ws.cell(row=i, column=5, value=r["prezzo_attuale"]).number_format = PRICE_FORMAT
        d_eur = ws.cell(row=i, column=6, value=r["variazione_eur"])
        d_eur.number_format = PRICE_FORMAT
        d_pct = ws.cell(row=i, column=7, value=round(r["variazione_pct"] / 100.0, 4))
        d_pct.number_format = "0.00%"
        ws.cell(row=i, column=8, value=_fmt_dt(r["data_prima"]))
        ws.cell(row=i, column=9, value=_fmt_dt(r["data_ultima"]))
        link = ws.cell(row=i, column=10, value=f'=HYPERLINK("{url}","Apri annuncio")')
        link.font = LINK_FONT

        # Convention from the spec:
        # * positive variation (rincaro) -> rosso
        # * negative variation (ribasso) -> verde (occasione)
        if r["variazione_eur"] is not None:
            color_font = (
                POSITIVE_FONT if r["variazione_eur"] > 0 else NEGATIVE_FONT
            )
            for col in (6, 7):
                ws.cell(row=i, column=col).font = color_font
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            if cell.font.size is None:
                cell.font = BODY_FONT

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    ws.freeze_panes = "A2"
    widths = [12, 18, 60, 14, 14, 14, 12, 18, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------------- "Riepilogo" sheet


def _build_riepilogo_sheet(
    wb: Workbook,
    ws,
    *,
    ctx: ExportContext,
    active_count: int,
    listings_rows: list[sqlite3.Row],
    n_listing_rows: int,
) -> None:
    durata_min = max(
        0,
        int((ctx.run_finished_at - ctx.run_started_at).total_seconds() // 60),
    )

    cards = [
        ("Nuovi annunci trovati", ctx.run_counters.get("n_new", 0)),
        ("Annunci con prezzo aggiornato", ctx.run_counters.get("n_updated", 0)),
        ("Annunci confermati (invariati)", ctx.run_counters.get("n_unchanged", 0)),
        ("Annunci scomparsi (non più disponibili)", ctx.run_counters.get("n_stale", 0)),
        ("Totale attivi nel DB", active_count),
        ("Durata run (minuti)", durata_min),
    ]

    ws["A1"] = "Riepilogo del run"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A2"] = ctx.run_started_at.strftime("%Y-%m-%d %H:%M") + " → " + ctx.run_finished_at.strftime("%Y-%m-%d %H:%M")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    for i, (label, value) in enumerate(cards, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(
            name="Calibri", size=11, bold=True
        )
        ws.cell(row=i, column=2, value=value).font = Font(name="Calibri", size=11)

    # Distribution by province (table + bar chart).
    by_prov = _aggregate(listings_rows, "provincia")
    by_stato = _aggregate(listings_rows, "stato")

    start_row_prov = 4 + len(cards) + 2
    ws.cell(row=start_row_prov, column=1, value="Distribuzione per provincia").font = Font(
        name="Calibri", size=12, bold=True
    )
    ws.cell(row=start_row_prov + 1, column=1, value="Provincia").font = HEADER_FONT
    ws.cell(row=start_row_prov + 1, column=1).fill = HEADER_FILL
    ws.cell(row=start_row_prov + 1, column=2, value="Annunci").font = HEADER_FONT
    ws.cell(row=start_row_prov + 1, column=2).fill = HEADER_FILL

    for offset, (key, count) in enumerate(by_prov, start=1):
        ws.cell(row=start_row_prov + 1 + offset, column=1, value=key or "—")
        ws.cell(row=start_row_prov + 1 + offset, column=2, value=count)

    if by_prov:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Annunci per provincia"
        chart.y_axis.title = "Provincia"
        chart.x_axis.title = "# annunci"
        data = Reference(
            ws,
            min_col=2,
            min_row=start_row_prov + 1,
            max_row=start_row_prov + 1 + len(by_prov),
        )
        cats = Reference(
            ws,
            min_col=1,
            min_row=start_row_prov + 2,
            max_row=start_row_prov + 1 + len(by_prov),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 16
        ws.add_chart(chart, f"D{start_row_prov}")

    # Distribution by stato.
    start_row_stato = start_row_prov + len(by_prov) + 4
    ws.cell(row=start_row_stato, column=1, value="Distribuzione per stato").font = Font(
        name="Calibri", size=12, bold=True
    )
    ws.cell(row=start_row_stato + 1, column=1, value="Stato").font = HEADER_FONT
    ws.cell(row=start_row_stato + 1, column=1).fill = HEADER_FILL
    ws.cell(row=start_row_stato + 1, column=2, value="Annunci").font = HEADER_FONT
    ws.cell(row=start_row_stato + 1, column=2).fill = HEADER_FILL
    for offset, (key, count) in enumerate(by_stato, start=1):
        ws.cell(row=start_row_stato + 1 + offset, column=1, value=key or "—")
        ws.cell(row=start_row_stato + 1 + offset, column=2, value=count)

    # Statistiche prezzi: formule che riferiscono il foglio Listing per
    # restare aggiornate quando l'utente modifica le righe.
    start_row_stats = start_row_stato + len(by_stato) + 4
    ws.cell(row=start_row_stats, column=1, value="Statistiche prezzi (€)").font = Font(
        name="Calibri", size=12, bold=True
    )
    if n_listing_rows >= 1:
        price_col = get_column_letter(_COL["Prezzo (€)"])
        rng = f"Listing!{price_col}2:{price_col}{n_listing_rows + 1}"
        stats = [
            ("Min", f"=MIN({rng})"),
            ("Q1", f"=QUARTILE({rng},1)"),
            ("Mediana", f"=MEDIAN({rng})"),
            ("Q3", f"=QUARTILE({rng},3)"),
            ("Max", f"=MAX({rng})"),
            ("Media", f"=AVERAGE({rng})"),
        ]
        for offset, (label, formula) in enumerate(stats, start=1):
            ws.cell(row=start_row_stats + offset, column=1, value=label).font = Font(
                name="Calibri", size=11, bold=True
            )
            cell = ws.cell(row=start_row_stats + offset, column=2, value=formula)
            cell.number_format = PRICE_FORMAT
    else:
        ws.cell(row=start_row_stats + 1, column=1, value="—")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24


def _aggregate(rows: list[sqlite3.Row], key: str) -> list[tuple[str, int]]:
    counts: dict[str, int] = {}
    for r in rows:
        k = r[key] if key in r.keys() else None
        counts[k or ""] = counts.get(k or "", 0) + 1
    return sorted(counts.items(), key=lambda t: t[1], reverse=True)


# ------------------------------------------------- "Run history" sheet


def _build_run_history_sheet(ws, runs: list[sqlite3.Row]) -> None:
    headers = [
        "Run #", "Inizio", "Fine", "Durata (sec)",
        "Nuovi", "Aggiornati", "Invariati", "Spariti",
    ]
    _write_header(ws, headers)

    for i, r in enumerate(runs, start=2):
        ws.cell(row=i, column=1, value=int(r["id"]))
        ws.cell(row=i, column=2, value=_fmt_dt(r["started_at"]))
        ws.cell(row=i, column=3, value=_fmt_dt(r["finished_at"]))
        ws.cell(row=i, column=4, value=r["durata_sec"])
        ws.cell(row=i, column=5, value=r["n_new"])
        ws.cell(row=i, column=6, value=r["n_updated"])
        ws.cell(row=i, column=7, value=r["n_unchanged"])
        ws.cell(row=i, column=8, value=r["n_stale"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).font = BODY_FONT

    if runs:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{len(runs) + 1}"
        )
    ws.freeze_panes = "A2"
    widths = [8, 20, 20, 14, 10, 12, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------- "Configurazione" sheet


def _build_config_sheet(ws, ctx: ExportContext) -> None:
    ws["A1"] = "Configurazione YAML del run"
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)

    yaml_lines = (ctx.config_yaml or "").splitlines() or ["(vuoto)"]
    for i, line in enumerate(yaml_lines, start=2):
        ws.cell(row=i, column=1, value=line).font = Font(
            name="Consolas", size=10
        )

    err_start = len(yaml_lines) + 4
    ws.cell(row=err_start, column=1, value="Errori e skip").font = Font(
        name="Calibri", size=12, bold=True
    )
    if not ctx.errors:
        ws.cell(row=err_start + 1, column=1, value="(nessuno)").font = Font(
            name="Calibri", size=10, italic=True
        )
    else:
        headers = ["Timestamp", "URL", "Tipo", "Messaggio"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=err_start + 1, column=col_idx, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        for i, err in enumerate(ctx.errors, start=err_start + 2):
            ws.cell(row=i, column=1, value=err.get("ts", ""))
            ws.cell(row=i, column=2, value=err.get("url", ""))
            ws.cell(row=i, column=3, value=err.get("kind", ""))
            ws.cell(row=i, column=4, value=err.get("message", ""))

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 60


# ------------------------------------------------------------ helpers


def _fmt_dt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    s = str(value)
    # Truncate to minute precision when the DB stores ISO seconds.
    if len(s) >= 16 and s[10] in (" ", "T"):
        return s[:16].replace("T", " ")
    return s
