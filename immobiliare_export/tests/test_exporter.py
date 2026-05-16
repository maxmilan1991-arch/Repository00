"""Tests for the xlsx exporter."""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from immobiliare_export.database import Database
from immobiliare_export.exporter import ExportContext, export_workbook
from immobiliare_export.models import Listing


def _mk_listing(**kw) -> Listing:
    defaults = dict(
        id=1, titolo="Casa di esempio", tipologia="Appartamento",
        regione="Lombardia", provincia="MI", comune="Milano",
        indirizzo="Via Roma 1", lat=45.46, lng=9.19,
        superficie_raw="120 m²", superficie_mq=120,
        stato="Buono / Abitabile", prezzo_corrente=300000,
        e_asta=False, e_trattativa=False, descrizione="…",
        locali=4, bagni=2, ricerca_nome="Milano", raw_json="{}",
    )
    defaults.update(kw)
    return Listing(**defaults)


@pytest.fixture
def db_with_data(tmp_path: Path) -> Database:
    db = Database(tmp_path / "exp.db")
    db.start_run("output_dir: ./out")  # pre-existing run for "Run history"
    base = datetime(2024, 1, 1, 10, 0, 0)
    db.upsert_listing(_mk_listing(id=1, prezzo_corrente=300000), seen_at=base)
    db.upsert_listing(_mk_listing(id=2, prezzo_corrente=500000,
                                  stato="Da ristrutturare"), seen_at=base)
    db.upsert_listing(_mk_listing(id=3, prezzo_corrente=200000,
                                  e_asta=True, titolo="Casa all'asta"),
                       seen_at=base)
    # Listing #1 had a price drop later.
    db.upsert_listing(_mk_listing(id=1, prezzo_corrente=280000),
                      seen_at=base + timedelta(days=10))
    yield db
    db.close()


def test_export_creates_xlsx_with_all_sheets(tmp_path: Path, db_with_data):
    out = tmp_path / "report.xlsx"
    ctx = ExportContext(
        config_yaml="output_dir: ./out\nsearches:\n  - nome: x\n    url: https://x",
        run_started_at=datetime(2024, 1, 5, 10, 0, 0),
        run_finished_at=datetime(2024, 1, 5, 10, 30, 0),
        run_counters={"n_new": 3, "n_updated": 1, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Listing", "Novità", "Variazioni di prezzo",
        "Riepilogo", "Run history", "Configurazione",
    ]


def test_listing_sheet_has_expected_rows(tmp_path: Path, db_with_data):
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime.utcnow(),
        run_finished_at=datetime.utcnow(),
        run_counters={"n_new": 0, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    wb = load_workbook(out)
    ws = wb["Listing"]
    # Header + 3 active listings
    assert ws.max_row == 4
    headers = [c.value for c in ws[1]]
    assert headers[0] == "#"
    assert "Prezzo (€)" in headers
    assert "Superficie (m²)" in headers
    assert "Link" in headers


def _find_col(ws, header_name: str) -> int:
    headers = [c.value for c in ws[1]]
    return headers.index(header_name) + 1


def test_surface_column_is_numeric(tmp_path: Path, db_with_data):
    """The "Superficie (m²)" column must contain integers, not strings,
    so Excel can sort/filter the column numerically. Seed listings have
    superficie_mq=120."""
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime.utcnow(),
        run_finished_at=datetime.utcnow(),
        run_counters={"n_new": 0, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    wb = load_workbook(out)
    ws = wb["Listing"]
    surf_col = _find_col(ws, "Superficie (m²)")
    cell = ws.cell(row=2, column=surf_col)
    assert cell.value == 120
    assert isinstance(cell.value, int)
    assert cell.number_format == "#,##0"


def test_surface_column_blank_when_unparseable(tmp_path: Path):
    """When superficie_mq is NULL (e.g. raw was 'n.d.'), the cell must
    be empty, not the literal 'None' or the raw string."""
    db = Database(tmp_path / "blank.db")
    db.upsert_listing(_mk_listing(id=1, superficie_raw="n.d.", superficie_mq=None))
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime.utcnow(),
        run_finished_at=datetime.utcnow(),
        run_counters={"n_new": 1, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db, out, ctx)

    wb = load_workbook(out)
    ws = wb["Listing"]
    surf_col = _find_col(ws, "Superficie (m²)")
    assert ws.cell(row=2, column=surf_col).value is None
    db.close()


def test_novita_surface_column_is_numeric(tmp_path: Path, db_with_data):
    """The Novità sheet reuses the Listing layout — same numeric rule."""
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime(2023, 1, 1, 10, 0, 0),  # seed is 2024-01-01
        run_finished_at=datetime(2023, 1, 1, 10, 30, 0),
        run_counters={"n_new": 3, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    wb = load_workbook(out)
    ws = wb["Novità"]
    surf_col = _find_col(ws, "Superficie (m²)")
    cell = ws.cell(row=2, column=surf_col)
    assert isinstance(cell.value, int)
    assert cell.value == 120
    assert cell.number_format == "#,##0"


def test_novita_sheet_empty_message(tmp_path: Path, db_with_data):
    """If no listings have first_seen >= run_started_at, show a friendly note."""
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        # All test data was seeded at 2024-01-01; setting run_started in the
        # future guarantees no listing qualifies as a novità.
        run_started_at=datetime(2030, 1, 1, 10, 0, 0),
        run_finished_at=datetime(2030, 1, 1, 10, 30, 0),
        run_counters={"n_new": 0, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    wb = load_workbook(out)
    ws = wb["Novità"]
    assert ws["A1"].value == "Nessuna novità in questo run"


def test_novita_sheet_lists_new_entries(tmp_path: Path, db_with_data):
    """If the run started before our seed timestamp, all 3 listings appear."""
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime(2023, 1, 1, 10, 0, 0),
        run_finished_at=datetime(2023, 1, 1, 10, 30, 0),
        run_counters={"n_new": 3, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)
    wb = load_workbook(out)
    ws = wb["Novità"]
    # Header + 3 entries
    assert ws.max_row == 4


def test_variazioni_sheet_shows_price_changes(tmp_path: Path, db_with_data):
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime.utcnow(),
        run_finished_at=datetime.utcnow(),
        run_counters={"n_new": 0, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    wb = load_workbook(out)
    ws = wb["Variazioni di prezzo"]
    # Only listing #1 had a price change → header + 1 row.
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=4).value == 300000  # initial price
    assert ws.cell(row=2, column=5).value == 280000  # current price


def test_variazioni_sheet_empty_when_no_changes(tmp_path: Path):
    db = Database(tmp_path / "empty.db")
    db.upsert_listing(_mk_listing(id=99))  # one listing, never updated
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime.utcnow(),
        run_finished_at=datetime.utcnow(),
        run_counters={"n_new": 1, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db, out, ctx)
    wb = load_workbook(out)
    ws = wb["Variazioni di prezzo"]
    assert ws["A2"].value == "Nessuna variazione di prezzo registrata"
    db.close()


def test_link_column_uses_hyperlink_formula(tmp_path: Path, db_with_data):
    out = tmp_path / "r.xlsx"
    ctx = ExportContext(
        config_yaml="",
        run_started_at=datetime.utcnow(),
        run_finished_at=datetime.utcnow(),
        run_counters={"n_new": 0, "n_updated": 0, "n_unchanged": 0, "n_stale": 0},
    )
    export_workbook(db_with_data, out, ctx)

    wb = load_workbook(out)
    ws = wb["Listing"]
    # Find the Link column.
    headers = [c.value for c in ws[1]]
    link_col = headers.index("Link") + 1
    formula = ws.cell(row=2, column=link_col).value
    assert isinstance(formula, str) and formula.startswith("=HYPERLINK(")
    assert "immobiliare.it/annunci/" in formula
